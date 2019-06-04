from bs4 import BeautifulSoup
from selenium import webdriver
import time
from openpyxl import load_workbook
from openpyxl import Workbook
import os
def save_excel(_FILENAME, _DATA, _HEADER):
    if os.path.exists(_FILENAME):
        if _DATA == None:
            return None
        book = load_workbook(_FILENAME)
        sheet = book.active
        for data in _DATA:
            sheet.append(data)
        book.save(_FILENAME)
    else:  # 새로만드는건
        if _HEADER == None:
            print(">>> 헤더 리스트를 먼저 넣어주세요")
            return None
        book = Workbook()
        sheet = book.active
        sheet.title = 'result'
        sheet.append(_HEADER)
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 40
        sheet.column_dimensions['E'].width = 40
        book.save(_FILENAME)
''' CONFIG AREA '''
# modify the value
DRIVER_PATH = "./chromedriver"
URL_DIR = "link/"
FILE_NAME = "한겨레_암_20181215-20181231_URL" # without extension
headerList = ['paper', 'date', 'title', 'article', 'web.aadr']
save_excel(FILE_NAME+'.xlsx', None, headerList) # init

dataList = []
browser = webdriver.Chrome(DRIVER_PATH)
f = open(URL_DIR+FILE_NAME+'.txt', 'r')
links = f.readlines()
f.close()

if __name__ == "__main__":
    for link in links:
        try:
            link = link.strip()
            print(">>", link, end=' : ')
            browser.get(link)
            time.sleep(5)  # 페이지 글 로딩되는거 기다려주기
            # 페이지 소스 얻어오기
            bs4 = BeautifulSoup(browser.page_source, 'lxml')

            if "well" in link or "ecotopia" in link:
                # 제목 가져오기
                title = bs4.find('div', attrs={'id': 'contentBody'}).find('h1').get_text()
                # 날짜 가져오기
                date = bs4.find('span', attrs={'class': 'date'}).get_text()
                # 내용 가져오기
                for tag in bs4.find_all('span', 'image_link_wrap'):
                    tag.extract()
                article = bs4.find('div', attrs={'class': 'xe_content'}).get_text()

            else:
                # 제목 가져오기
                title = bs4.find('span', attrs={'class': 'title'}).get_text()
                # 날짜 가져오기
                date = bs4.find('p',class_='date-time').find('span').get_text()
                date = date.replace("등록 :", "").lstrip()
                # 내용 가져오기
                for tag in bs4.find_all('div', 'desc'):
                    tag.extract()
                for tag in bs4.find_all('script'):
                    tag.extract()
                for tag in bs4.find_all('div', 'image-area'):
                    tag.extract()
                article = bs4.find('div', attrs={'class': 'text'}).get_text()

            dataList.append(["한겨레", date, title, article, link])
            print("완료")
            if len(dataList) >= 5:
                save_excel(FILE_NAME + '.xlsx', dataList, None)
                dataList.clear()
                dataList = []
        except:
            print("완료.")
            pass
    
    save_excel(FILE_NAME + '.xlsx', dataList, None)
    browser.quit()

