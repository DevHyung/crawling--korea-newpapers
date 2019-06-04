from bs4 import BeautifulSoup
from selenium import webdriver
import time
from openpyxl import load_workbook
from openpyxl import Workbook
import os
def save_excel(_FILENAME, _DATA, _HEADER):
    if os.path.exists(_FILENAME):
        if _DATA == None:
            return None
        book = load_workbook(_FILENAME)
        sheet = book.active
        for data in _DATA:
            sheet.append(data)
        book.save(_FILENAME)
    else:  # 새로만드는건
        if _HEADER == None:
            print(">>> 헤더 리스트를 먼저 넣어주세요")
            return None
        book = Workbook()
        sheet = book.active
        sheet.title = 'result'
        sheet.append(_HEADER)
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 40
        sheet.column_dimensions['E'].width = 40
        book.save(_FILENAME)
''' CONFIG AREA '''
# modify the value
DRIVER_PATH = "./chromedriver"
URL_DIR = "link/"
FILE_NAME = "조선일보_20181215-20181231_URL" # without extension
headerList = ['paper', 'date', 'title', 'article', 'web.aadr']
save_excel(FILE_NAME+'.xlsx', None, headerList) # init

dataList = []
browser = webdriver.Chrome(DRIVER_PATH)
f = open(URL_DIR+FILE_NAME+'.txt', 'r')
links = f.readlines()
f.close()

if __name__ == "__main__":
    for link in links:
        result = ""
        try:
            link = link.strip()
            print(">>", link, end=' : ')
            if "biz" in link:
                browser.get(link)
                time.sleep(5)  # 페이지 글 로딩되는거 기다려주기
                bs4 = BeautifulSoup(browser.page_source, 'lxml')
                title = bs4.find('h1', attrs={'id': 'news_title_text_id'}).get_text()
                date = bs4.find('div', attrs={'class': 'news_date'}).get_text()
                date = date.replace("입력", "").split('|', 1)[0].lstrip()
                articles = bs4.find_all('div', attrs={'class': 'par'})
                for article in articles:
                    for span in bs4.find_all("span", attrs={'class': 'par_in_ad'}):
                        span.decompose()
                    result += article.get_text().strip()
            else:
                browser.get(link)
                time.sleep(5)  # 페이지 글 로딩되는거 기다려주기
                bs4 = BeautifulSoup(browser.page_source, 'lxml')
                title = bs4.find('h1', attrs={'id': 'news_title_text_id'}).get_text()
                date = bs4.find('div', attrs={'class': 'news_date'}).get_text()
                date = date.replace("입력", "").split('|', 1)[0].lstrip()
                articles = bs4.find_all('div', attrs={'class': 'par'})
                for article in articles:
                    for span in bs4.find_all("span", attrs={'class': 'par_in_ad'}):
                        span.decompose()
                    result += article.get_text().strip()
            dataList.append(["조선일보", date, title, result, link])
            print("완료")
            if len(dataList) >= 5:
                save_excel(FILE_NAME + '.xlsx', dataList, None)
                dataList.clear()
                dataList = []
        except:
            print("완료.")
            pass
    save_excel(FILE_NAME + '.xlsx', dataList, None)
    browser.quit()
