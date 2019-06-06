from bs4 import BeautifulSoup
from selenium import webdriver
import time
from openpyxl import load_workbook
from openpyxl import Workbook
import os
def save_excel(_FILENAME, _DATA, _HEADER):
    if os.path.exists(_FILENAME):
        if _DATA == None:
            return None
        book = load_workbook(_FILENAME)
        sheet = book.active
        for data in _DATA:
            sheet.append(data)
        book.save(_FILENAME)
    else:  # 새로만드는건
        if _HEADER == None:
            print(">>> 헤더 리스트를 먼저 넣어주세요")
            return None
        book = Workbook()
        sheet = book.active
        sheet.title = 'result'
        sheet.append(_HEADER)
        sheet.column_dimensions['A'].width = 10
        sheet.column_dimensions['B'].width = 15
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 40
        sheet.column_dimensions['E'].width = 40
        book.save(_FILENAME)
''' CONFIG AREA '''
# modify the value
DRIVER_PATH = "./chromedriver"
URL_DIR = "link/"
FILE_NAME = "경향신문_암_20090101-20181231_URL" # without extension
headerList = ['paper', 'date', 'title', 'article', 'web.aadr']
save_excel(FILE_NAME+'.xlsx', None, headerList) # init

dataList = []
browser = webdriver.Chrome(DRIVER_PATH)
f = open(URL_DIR+FILE_NAME+'.txt', 'r')
links = f.readlines()
f.close()

if __name__ == "__main__":
    for link in links:
        result = ""
        try:
            link = link.strip()
            print(">>", link, end=' : ')
            browser.get(link)
            time.sleep(3)  # 페이지 글 로딩되는거 기다려주기
            bs4 = BeautifulSoup(browser.page_source, 'lxml')
            if "h2" in browser.current_url:
                title = bs4.find('div', attrs={'class': 'art_tit'}).get_text().strip()
                dates = bs4.find_all('div', attrs={'class': 'art_date'})
                for d in dates:
                    if "수정" not in d.get_text():
                        date = d.get_text()
                date = date.replace("입력", "").strip()
                for tag in bs4.find_all('a', attrs={'target': '_blank'}):
                    tag.extract()
                articles = bs4.find_all('p', attrs={'class': 'art_text'})
                for article in articles:
                    result += article.get_text().strip()

            elif "biz" in browser.current_url:
                title = bs4.find('h1', attrs={'id': 'articleTtitle'}).get_text().strip()
                # 날짜 가져오기
                #date = bs4.select('div.byline>em')[0].text
                date = bs4.find('div',id='bylineArea').find('em').get_text().strip()
                date = date.replace("입력 :", "").lstrip()
                articles = bs4.find_all('p', attrs={'class': 'content_text'})
                for article in articles:
                    result += article.get_text().strip()
            else:
                title = bs4.find('h1', attrs={'id': 'article_title'}).get_text().strip()
                date = bs4.find('div',class_='byline').get_text()
                date = date.replace("입력 :", "").lstrip()
                # 내용 가져오기
                for tag in bs4.find_all('div', 'art_photo_wrap'):
                    tag.extract()
                article = bs4.find('div', attrs={'class': 'art_body'}).get_text().strip()
                result = article
            dataList.append(["경향신문", date, title, result, link])
            print("완료")

            while True: #팝업창 닫기
                if len(browser.window_handles) >= 2:
                    main_window_handle = browser.current_window_handle

                    signin_window_handle = None
                    while not signin_window_handle:
                        for handle in browser.window_handles:
                            if handle != main_window_handle:
                                signin_window_handle = handle
                                break
                    browser.switch_to.window(signin_window_handle)
                    browser.close()
                    browser.switch_to.window(main_window_handle)
                else:
                    break
            if len(dataList) >= 5:
                save_excel(FILE_NAME + '.xlsx', dataList, None)
                dataList.clear()
                dataList = []
        except:
            print("완료.")
            pass
    save_excel(FILE_NAME + '.xlsx', dataList, None)
    browser.quit()